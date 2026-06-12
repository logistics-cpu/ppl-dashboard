"""
One-time seeding of the product-cost tables from "📦 Product Cost 2026.xlsx".

Parses the workbook's tabs into the cost_* tables:
  'US '                      → cost_products (manual inputs + frozen overrides)
  'MasterData ' + 'SKU Master' + 'SKU Rent Calculation' → cost_sku_specs
  '📦 Shipment Data'          → cost_shipments + cost_shipment_lines
  'Warehouse Rent Rate Table' + 'Rate Card' + 'Settings' → rate tables
  'Classification'           → cost_lastmile_orders (order-level history)

Parsing rules that mirror the workbook's quirks:
  - US sheet: header row 2, data row 3+. Section-title rows (name only,
    no costs) become the `category` for following rows.
  - Computed columns G-K: if the cell is a VLOOKUP formula the value is
    left to the dashboard to recompute; if it's a literal number or a
    cross-row formula (composites), the workbook's computed value is
    frozen into the matching *_override column.
  - Multi-SKU cells (newline-separated, optional *N suffix) split into
    one cost_product row per SKU sharing the same costs.
  - Cozy-bundle rows reference the global Type-C last-mile average →
    lastmile_group='TYPE_C'. Combo-cover rows look up by bundle name →
    lastmile_group='COVERPAIR:<color>' so they keep auto-updating.
"""

import re
from io import BytesIO

import openpyxl

from core.costs import (
    REGION_US, _u, classify_order,
    replace_cost_products, replace_sku_specs, replace_shipments,
    replace_rent_brackets, replace_rate_card,
    insert_lastmile_orders_bulk, take_snapshot,
)
from core.database import get_db, set_setting, invalidate_cache


def _num(v):
    """Coerce workbook cell to float (handles text-typed numbers), else None."""
    if v is None or v == "":
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _is_formula(cell):
    return isinstance(cell.value, str) and cell.value.startswith("=")


def _split_sku_cell(value):
    """Split a newline-separated SKU cell; strip '*N' multipliers; dedupe."""
    if not value:
        return []
    out = []
    for line in str(value).split("\n"):
        s = re.sub(r"\*\d+\s*$", "", line.strip()).strip()
        if s and s not in out:
            out.append(s)
    return out


def _coverpair_group_for_name(name):
    """Map a combo-cover product name to its COVERPAIR color token."""
    n = (name or "").lower()
    if "ice" in n:
        return "COVERPAIR:ICE"
    if "cream" in n:
        return "COVERPAIR:NEWYELLOW"
    if "turquoise" in n:
        return "COVERPAIR:NEWBLUE"
    if "peach" in n:
        return "COVERPAIR:NEWPINK"
    return None


# ===========================================================================
# US sheet → cost_products
# ===========================================================================

def _parse_bundle_names(wbv):
    """
    'Bundle Names' tab: sorted SKU key → custom bundle name. Inverted here
    to {NORMALIZED_NAME: set of non-accessory product SKUs} so legging rows
    can pick up old/new SKU variants the US sheet doesn't list.
    """
    from core.costs import BAG_SKU, LARGEBOX_SKU
    name_map = {}
    try:
        ws = wbv["Bundle Names"]
    except KeyError:
        return name_map
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[0] or not row[1]:
            continue
        name = " ".join(str(row[1]).split()).upper()
        skus = {
            _u(s) for s in str(row[0]).split(";")
            if s.strip() and _u(s) not in (BAG_SKU, LARGEBOX_SKU)
        }
        name_map.setdefault(name, set()).update(skus)
    return name_map


def parse_us_sheet(wbv, wbf, region=REGION_US):
    """
    Returns (rows, warnings): rows ready for replace_cost_products.
    wbv = values workbook (data_only=True), wbf = formulas workbook.
    """
    wsv, wsf = wbv["US "], wbf["US "]
    rows, warnings = [], []
    category = None
    seen_keys = set()
    bundle_names = _parse_bundle_names(wbv)

    for r in range(3, wsv.max_row + 1):
        vals = [wsv.cell(row=r, column=c).value for c in range(1, 17)]
        if not any(v is not None and v != "" for v in vals):
            continue
        a, b, c_cell, name = vals[0], vals[1], vals[2], vals[3]
        product_cost = _num(vals[4])

        # Section title: a name with no SKUs and no product cost
        if name and not a and not b and not c_cell and product_cost is None:
            category = str(name).strip()
            continue
        if product_cost is None and not a and not c_cell:
            continue  # stray row

        agent_fee = _num(vals[5])
        pick_pack = _num(vals[11])
        pink_box = _num(vals[12])
        other_box = _num(vals[13])

        # Computed columns G..K (cols 7-11): formula → recompute in dashboard;
        # literal / cross-row formula → freeze the workbook value as override.
        overrides = {}
        computed_cols = [
            (7, "domestic_override"), (8, "sea_override"), (9, "rent_override"),
            (10, "inbound_override"), (11, "lastmile_override"),
        ]
        is_vlookup = {}
        for col, key in computed_cols:
            fcell = wsf.cell(row=r, column=col)
            vval = _num(wsv.cell(row=r, column=col).value)
            if _is_formula(fcell) and "VLOOKUP" in str(fcell.value):
                is_vlookup[key] = True          # dashboard recomputes
            elif fcell.value is None:
                is_vlookup[key] = True          # empty → recompute (will flag)
            else:
                is_vlookup[key] = False
                overrides[key] = vval           # literal or cross-row formula

        is_composite = bool(a and str(a).startswith("No SKU")) or (not a and not b)

        china1_list = _split_sku_cell(a) if not str(a or "").startswith("No SKU") else []
        china2_list = _split_sku_cell(b)
        china_sku1 = _u(china1_list[0]) if china1_list else None
        china_sku2 = _u(china2_list[0]) if china2_list else None

        # last-mile group
        lastmile_group = "SINGLE"
        k_formula = wsf.cell(row=r, column=11).value
        if isinstance(k_formula, str) and k_formula.startswith("="):
            if "VLOOKUP" not in k_formula:
                # direct cell ref → the global Type-C (cozy) average
                lastmile_group = "TYPE_C"
                overrides.pop("lastmile_override", None)
            elif f"VLOOKUP(D{r}" in k_formula.replace(" ", ""):
                # lookup by product NAME (a bundle-name average):
                #  - combo covers → color-specific cover-pair pooling
                #  - leggings (ship with the packaging bag) → WITHBAG pooling
                grp = _coverpair_group_for_name(name)
                if grp and ("cover" in str(name).lower()):
                    lastmile_group = grp
                else:
                    lastmile_group = "WITHBAG"
                    # Pick up SKU variants from Bundle Names that the US
                    # row doesn't list (e.g. NEWBLACK old/new pairs), so
                    # the pooled type-B average sees all variants' orders.
                    norm_name = " ".join(str(name or "").split()).upper()
                    siblings = bundle_names.get(norm_name, set())
                    listed = {china_sku1, china_sku2} - {None}
                    extra = sorted(siblings - listed)
                    if extra and not china_sku2:
                        china_sku2 = extra[0]
                overrides.pop("lastmile_override", None)

        # Shopify SKU(s): prefer col C; fall back to the China SKU lines so
        # products without a Shopify mapping (hydration, new colors) still
        # get one row each.
        shopify_lines = _split_sku_cell(c_cell)
        if not shopify_lines:
            shopify_lines = china1_list or [None]

        for idx, sline in enumerate(shopify_lines):
            shop_upper = _u(sline)
            if not shop_upper:
                # Rows with no SKU at all (reference bundles) get a
                # deterministic slug key so they still display and edit.
                slug = re.sub(r"[^A-Z0-9]+", "-", str(name or f"ROW{r}").upper()).strip("-")
                shop_upper = f"BUNDLE-{slug[:50]}"
                sline = shop_upper
            if shop_upper in seen_keys:
                warnings.append(f"row {r}: duplicate shopify SKU '{sline}' skipped")
                continue
            seen_keys.add(shop_upper)

            # Per-SKU rows split from a grouped China-SKU cell (e.g. the
            # 136181 color leggings) get their own China SKU for lookups.
            c1, c2 = china_sku1, china_sku2
            if not _split_sku_cell(c_cell) and len(china1_list) > 1:
                c1, c2 = _u(china1_list[idx]) if idx < len(china1_list) else china_sku1, None

            rows.append({
                "region": region,
                "shopify_sku": shop_upper,
                "display_sku": sline,
                "product_name": str(name).strip() if name else None,
                "category": category,
                "china_sku1": c1,
                "china_sku2": c2,
                "is_composite": 1 if is_composite else 0,
                "product_cost": product_cost,
                "agent_fee": agent_fee,
                "pick_pack": pick_pack,
                "pink_box": pink_box,
                "other_box": other_box,
                "domestic_override": overrides.get("domestic_override"),
                "sea_override": overrides.get("sea_override"),
                "rent_override": overrides.get("rent_override"),
                "inbound_override": overrides.get("inbound_override"),
                "lastmile_override": overrides.get("lastmile_override"),
                "lastmile_group": lastmile_group,
                "notes": None,
                "active": 1,
            })

        # Grouped China-SKU rows with no Shopify SKU and >1 china line:
        # emit the remaining lines too (idx beyond shopify_lines length)
        if not _split_sku_cell(c_cell) and len(china1_list) > 1:
            for extra in china1_list[len(shopify_lines):]:
                shop_upper = _u(extra)
                if shop_upper in seen_keys:
                    continue
                seen_keys.add(shop_upper)
                base = rows[-1].copy()
                base["shopify_sku"] = shop_upper
                base["display_sku"] = extra
                base["china_sku1"] = shop_upper
                base["china_sku2"] = None
                rows.append(base)

    return rows, warnings


# ===========================================================================
# Specs (MasterData + SKU Master + SKU Rent Calculation)
# ===========================================================================

def parse_specs(wbv, region=REGION_US):
    specs = {}

    def entry(sku):
        k = _u(sku)
        if k not in specs:
            specs[k] = {
                "region": region, "sku": k,
                "unit_cbm": None, "unit_weight_kg": None,
                "qty_per_ctn": None, "cbm_per_ctn": None, "vol_weight_ctn": None,
                "rent_unit_cbm": None, "assumed_storage_days": 90,
                "in_sku_master": 0, "in_rent_table": 0,
            }
        return specs[k]

    # MasterData: SKU | Vol Weight/Ctn | Qty/Ctn | CBM/Ctn (header r3)
    ws = wbv["MasterData "]
    for row in ws.iter_rows(min_row=4, values_only=True):
        if not row[0]:
            continue
        e = entry(row[0])
        e["vol_weight_ctn"] = _num(row[1])
        e["qty_per_ctn"] = _num(row[2])
        e["cbm_per_ctn"] = _num(row[3])

    # SKU Master: SKU | Unit CBM | Unit Weight | ... (header r2)
    ws = wbv["SKU Master"]
    for row in ws.iter_rows(min_row=3, values_only=True):
        if not row[0]:
            continue
        e = entry(row[0])
        e["unit_cbm"] = _num(row[1])
        e["unit_weight_kg"] = _num(row[2])
        e["in_sku_master"] = 1

    # SKU Rent Calculation — right "general" block: G=SKU H=CBM I=days J=rent
    ws = wbv["SKU Rent Calculation"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        sku, cbm, days = (row[6] if len(row) > 6 else None), \
                         (row[7] if len(row) > 7 else None), \
                         (row[8] if len(row) > 8 else None)
        if not sku:
            continue
        e = entry(sku)
        rent_cbm = _num(cbm)
        if rent_cbm is not None:
            e["rent_unit_cbm"] = rent_cbm
        e["in_rent_table"] = 1
        if _num(days):
            e["assumed_storage_days"] = int(_num(days))

    # Fill unit_cbm from carton specs when SKU Master doesn't have the SKU
    for e in specs.values():
        if e["unit_cbm"] is None and e["cbm_per_ctn"] and e["qty_per_ctn"]:
            e["unit_cbm"] = e["cbm_per_ctn"] / e["qty_per_ctn"]

    return list(specs.values())


# ===========================================================================
# Freight shipments ('📦 Shipment Data', header r4, data r6+)
# ===========================================================================

def parse_shipments(wbv, region=REGION_US):
    ws = wbv["📦 Shipment Data"]
    headers, lines, warnings = {}, [], []
    for row in ws.iter_rows(min_row=6, values_only=True):
        sid_raw, sku, qty = row[0], row[1], _num(row[2])
        if sid_raw is None or not sku or not qty:
            continue
        sid = str(int(sid_raw)) if isinstance(sid_raw, float) else str(sid_raw).strip()
        dom, sea = _num(row[3]) or 0.0, _num(row[4]) or 0.0
        notes = row[13] if len(row) > 13 else None
        if sid not in headers:
            headers[sid] = {
                "region": region, "shipment_id": sid, "ship_date": None,
                "dom_total": dom, "sea_total": sea,
                "notes": str(notes).strip() if notes else None,
            }
        else:
            h = headers[sid]
            if abs(h["dom_total"] - dom) > 0.01 or abs(h["sea_total"] - sea) > 0.01:
                warnings.append(
                    f"shipment {sid}: inconsistent totals across lines "
                    f"(kept first: dom {h['dom_total']}, sea {h['sea_total']})"
                )
        lines.append({
            "region": region, "shipment_id": sid,
            "sku": _u(str(sku)), "qty": int(qty),
        })
    return list(headers.values()), lines, warnings


# ===========================================================================
# Rate tables
# ===========================================================================

def parse_rates(wbv, region=REGION_US):
    # Warehouse Rent Rate Table: Bracket | Start | End | Rate (header r1)
    brackets = []
    for row in wbv["Warehouse Rent Rate Table"].iter_rows(min_row=2, values_only=True):
        if row[1] is None and row[3] is None:
            continue
        end = _num(row[2])
        brackets.append({
            "region": region,
            "start_day": _num(row[1]) or 0,
            "end_day": None if (end is None or end >= 99999) else end,
            "rate_per_cbm_day": _num(row[3]) or 0,
        })

    # Rate Card: Tier start | Tier end | Op fee (header r3)
    tiers = []
    for row in wbv["Rate Card"].iter_rows(min_row=4, values_only=True):
        if row[0] is None and row[2] is None:
            continue
        fee = _num(row[2])
        if fee is None:
            continue
        tiers.append({
            "region": region,
            "tier_start_kg": _num(row[0]) or 0,
            "tier_end_kg": _num(row[1]),
            "op_fee": fee,
        })

    unload_rate = _num(wbv["Settings"].cell(row=2, column=2).value)
    return brackets, tiers, unload_rate


# ===========================================================================
# Classification → cost_lastmile_orders
# ===========================================================================

def parse_classification(wbv, region=REGION_US):
    ws = wbv["Classification"]
    rows, skipped = [], 0
    for row in ws.iter_rows(min_row=3, values_only=True):
        oid = row[0]
        if not oid:
            continue
        ship_date = row[1]
        if hasattr(ship_date, "date"):
            ship_date = ship_date.date().isoformat()
        elif isinstance(ship_date, str):
            ship_date = ship_date.strip()[:10]
        if not ship_date:
            skipped += 1
            continue
        cost = _num(row[4])
        if cost is None:
            skipped += 1
            continue
        qty_raw = str(row[6] or "")
        try:
            total_qty = sum(int(float(q)) for q in qty_raw.split(";") if q.strip())
        except ValueError:
            total_qty = None
        sku_key = _u(str(row[13])) if row[13] and str(row[13]) != "N/A" else None
        # Re-classify with our own rules rather than trusting the workbook
        # column — the workbook types the same SKU combos inconsistently,
        # and recurring uploads will use this classifier anyway.
        if sku_key:
            skus = set(s for s in sku_key.split(";") if s)
            otype = classify_order(skus)
        else:
            otype = row[8] if row[8] in ("A", "B", "C") else "OTHER"
        rows.append({
            "region": region,
            "order_id": str(oid).strip(),
            "ship_date": ship_date,
            "country": str(row[2]).strip() if row[2] else None,
            "shipping_cost": cost,
            "sku_key": sku_key,
            "main_sku": _u(str(row[10])) if row[10] and str(row[10]) != "N/A" else None,
            "order_type": otype,
            "total_qty": total_qty,
            "num_skus": int(_num(row[7]) or 0) or None,
        })
    return rows, skipped


# ===========================================================================
# Orchestrator
# ===========================================================================

def seed_from_workbook(file_bytes, region=REGION_US, progress=None):
    """
    Full seed from the workbook. file_bytes: bytes of the .xlsx.
    progress: optional callable(str) for status updates.
    Returns a stats dict.
    """
    def _p(msg):
        if progress:
            progress(msg)

    _p("Opening workbook (values)…")
    wbv = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    _p("Opening workbook (formulas)…")
    wbf = openpyxl.load_workbook(BytesIO(file_bytes), data_only=False)

    stats = {"warnings": []}

    _p("Parsing US cost sheet…")
    products, warns = parse_us_sheet(wbv, wbf, region)
    stats["warnings"].extend(warns)
    _p(f"→ {len(products)} products. Parsing specs…")
    specs = parse_specs(wbv, region)
    _p(f"→ {len(specs)} SKU specs. Parsing shipments…")
    headers, lines, warns = parse_shipments(wbv, region)
    stats["warnings"].extend(warns)
    _p(f"→ {len(headers)} shipments / {len(lines)} lines. Parsing rates…")
    brackets, tiers, unload_rate = parse_rates(wbv, region)
    _p("Parsing last-mile order history (Classification)…")
    lastmile, skipped = parse_classification(wbv, region)
    if skipped:
        stats["warnings"].append(f"{skipped} classification rows skipped (no date/cost)")

    _p(f"Writing {len(products)} products to DB…")
    stats["products"] = replace_cost_products(products, region)
    _p("Writing SKU specs…")
    stats["specs"] = replace_sku_specs(specs, region)
    _p("Writing shipments…")
    stats["shipments"], stats["shipment_lines"] = replace_shipments(headers, lines, region)
    _p("Writing rate tables…")
    stats["rent_brackets"] = replace_rent_brackets(brackets, region)
    stats["rate_tiers"] = replace_rate_card(tiers, region)
    if unload_rate:
        set_setting("cost_us_unload_rate_per_cbm", unload_rate)

    _p(f"Writing {len(lastmile)} last-mile orders (this is the big one)…")
    with get_db() as conn:
        conn.execute("DELETE FROM cost_lastmile_orders WHERE region = ?", (region,))
        conn.commit()
    stats["lastmile_orders"] = insert_lastmile_orders_bulk(lastmile, region)

    _p("Taking baseline cost snapshot…")
    stats["snapshot_rows"] = take_snapshot(region, reason="seed")
    invalidate_cache()
    _p("Done.")
    return stats


# ===========================================================================
# Parity check — computed vs workbook
# ===========================================================================

def parity_check(file_bytes, region=REGION_US, tolerance=0.01):
    """
    Compare assemble_cost_table() against the workbook's US-sheet
    Total (O) / Landed (P) values.

    The dashboard deliberately computes last-mile bundle averages
    differently from the workbook (order-level pooling across SKU
    variants instead of first-match / mean-of-means lookups), so each
    diff is attributed: if the row matches once last-mile is excluded,
    it's flagged as a method deviation rather than an error.

    Returns (diffs, matched) — diffs have a `kind` of
    'lastmile_method' or 'error'.
    """
    from core.costs import assemble_cost_table

    wbv = openpyxl.load_workbook(BytesIO(file_bytes), data_only=True)
    ws = wbv["US "]

    # Component columns whose dashboard calculation deliberately improves
    # on the workbook (order-level last-mile pooling; rent CBM fallback
    # when the workbook's rent row has no CBM). Diffs confined to these
    # are reported as 'method', everything else as 'error'.
    METHOD_COMPONENTS = {"warehouse_rent", "local_shipping"}
    COMPONENT_COLS = [
        ("domestic_freight", 7), ("sea_freight", 8), ("warehouse_rent", 9),
        ("inbound", 10), ("local_shipping", 11),
    ]

    expected = {}  # UPPER sku → {total, landed, components{}}
    for r in range(3, ws.max_row + 1):
        c_cell = ws.cell(row=r, column=3).value
        a_cell = ws.cell(row=r, column=1).value
        total = _num(ws.cell(row=r, column=15).value)
        if total is None:
            continue
        comps = {
            key: _num(ws.cell(row=r, column=col).value)
            for key, col in COMPONENT_COLS
        }
        keys = _split_sku_cell(c_cell)
        if not keys and a_cell and not str(a_cell).startswith("No SKU"):
            keys = _split_sku_cell(a_cell)
        for k in keys:
            expected[_u(k)] = {
                "total": total,
                "landed": _num(ws.cell(row=r, column=16).value),
                "components": comps,
            }

    table = assemble_cost_table(region)
    diffs, matched = [], 0
    for row in table:
        sku = row["shopify_sku"]
        if not sku or sku not in expected:
            continue
        exp = expected[sku]
        d_total = abs(row["total_cost"] - exp["total"])
        d_landed = (
            abs(row["landed_cost"] - exp["landed"])
            if exp["landed"] is not None else 0
        )
        if d_total <= tolerance and d_landed <= tolerance:
            matched += 1
            continue
        # Attribute the diff to components
        diff_comps = []
        for key, _col in COMPONENT_COLS:
            wb_val = exp["components"][key] or 0.0
            if abs(row[key] - wb_val) > tolerance:
                diff_comps.append(key)
        kind = (
            "method"
            if diff_comps and set(diff_comps) <= METHOD_COMPONENTS
            else "error"
        )
        diffs.append({
            "kind": kind,
            "shopify_sku": sku,
            "product_name": row["product_name"],
            "computed_total": round(row["total_cost"], 4),
            "workbook_total": exp["total"],
            "computed_landed": round(row["landed_cost"], 4),
            "workbook_landed": exp["landed"],
            "diff_components": ", ".join(diff_comps) or "(manual inputs)",
            "missing": ", ".join(row["missing"]),
        })
    return diffs, matched
